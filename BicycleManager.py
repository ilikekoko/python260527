import sys
import sqlite3
import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QSpinBox, QPushButton, QTableWidget, QTableWidgetItem,
    QMessageBox
)
from PyQt6.QtCore import Qt


class BicycleManager(QMainWindow):
    def __init__(self):
        super().__init__()
        self.init_db()
        self.init_ui()
        
    def init_db(self):
        """데이터베이스 초기화 및 데이터 생성"""
        self.conn = sqlite3.connect('bicycle.db')
        self.cursor = self.conn.cursor()
        
        # 테이블 생성 (없으면)
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS Bycle (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                price INTEGER NOT NULL,
                qty INTEGER NOT NULL
            )
        ''')
        self.conn.commit()
        
        # 데이터가 없으면 초기 데이터 100개 생성
        self.cursor.execute('SELECT COUNT(*) FROM Bycle')
        if self.cursor.fetchone()[0] == 0:
            self.insert_sample_data()
    
    def insert_sample_data(self):
        """샘플 자전거 데이터 100개 생성"""
        bicycle_types = [
            '로드바이크', '마운틴바이크', '하이브리드바이크', '크루저', '싱글속도',
            'BMX', '팻바이크', '전기자전거', '픽시', '접이식자전거',
            '아동용자전거', '여성용자전거', '스포츠자전거', '출퇴근용자전거', 'MTB'
        ]
        
        colors = ['검정색', '흰색', '빨간색', '파란색', '녹색', '노란색', 
                  '주황색', '은색', '회색', '갈색']
        
        data = []
        for i in range(1, 101):
            bike_type = random.choice(bicycle_types)
            color = random.choice(colors)
            name = f"{bike_type} - {color}"
            price = random.randint(150000, 3000000)
            qty = random.randint(1, 50)
            data.append((name, price, qty))
        
        self.cursor.executemany(
            'INSERT INTO Bycle (name, price, qty) VALUES (?, ?, ?)',
            data
        )
        self.conn.commit()
    
    def init_ui(self):
        """UI 초기화"""
        self.setWindowTitle('자전거 관리 시스템')
        self.setGeometry(100, 100, 1200, 700)
        self.apply_stylesheet()
        
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        main_layout = QVBoxLayout(central_widget)
        
        # 검색/입력 영역
        input_layout = QHBoxLayout()
        
        # ID 검색
        input_layout.addWidget(QLabel('ID:'))
        self.search_id = QLineEdit()
        self.search_id.setMaximumWidth(100)
        input_layout.addWidget(self.search_id)
        input_layout.addWidget(QPushButton('검색', clicked=self.search_by_id))
        
        input_layout.addSpacing(20)
        
        # 자전거 이름
        input_layout.addWidget(QLabel('자전거명:'))
        self.name_input = QLineEdit()
        self.name_input.setMaximumWidth(200)
        input_layout.addWidget(self.name_input)
        
        # 가격
        input_layout.addWidget(QLabel('가격:'))
        self.price_input = QSpinBox()
        self.price_input.setMaximumWidth(100)
        self.price_input.setRange(0, 10000000)
        input_layout.addWidget(self.price_input)
        
        # 수량
        input_layout.addWidget(QLabel('수량:'))
        self.qty_input = QSpinBox()
        self.qty_input.setMaximumWidth(100)
        self.qty_input.setRange(0, 1000)
        input_layout.addWidget(self.qty_input)
        
        input_layout.addStretch()
        main_layout.addLayout(input_layout)
        
        # 버튼 영역
        button_layout = QHBoxLayout()
        button_layout.addWidget(QPushButton('추가', clicked=self.add_bicycle))
        button_layout.addWidget(QPushButton('수정', clicked=self.update_bicycle))
        button_layout.addWidget(QPushButton('삭제', clicked=self.delete_bicycle))
        button_layout.addWidget(QPushButton('초기화', clicked=self.clear_inputs))
        button_layout.addWidget(QPushButton('전체보기', clicked=self.load_all_data))
        button_layout.addWidget(QPushButton('엑셀로 출력', clicked=self.export_to_excel))
        button_layout.addStretch()
        main_layout.addLayout(button_layout)
        
        # 테이블 영역
        self.table = QTableWidget()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(['ID', '자전거명', '가격', '수량'])
        self.table.setColumnWidth(0, 50)
        self.table.setColumnWidth(1, 400)
        self.table.setColumnWidth(2, 150)
        self.table.setColumnWidth(3, 150)
        self.table.itemSelectionChanged.connect(self.load_selected_row)
        main_layout.addWidget(QLabel('데이터 목록:'))
        main_layout.addWidget(self.table)
        
        # 초기 데이터 로드
        self.load_all_data()
    
    def load_all_data(self):
        """모든 데이터 로드"""
        self.cursor.execute('SELECT * FROM Bycle ORDER BY id')
        rows = self.cursor.fetchall()
        self.display_data(rows)
    
    def display_data(self, rows):
        """테이블에 데이터 표시"""
        self.table.setRowCount(len(rows))
        for row_idx, row in enumerate(rows):
            for col_idx, value in enumerate(row):
                item = QTableWidgetItem(str(value))
                if col_idx == 0:  # ID는 읽기 전용
                    item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                self.table.setItem(row_idx, col_idx, item)
    
    def load_selected_row(self):
        """선택된 행의 데이터를 입력 필드에 로드"""
        selected_rows = self.table.selectedIndexes()
        if not selected_rows:
            return
        
        row = selected_rows[0].row()
        self.search_id.setText(self.table.item(row, 0).text())
        self.name_input.setText(self.table.item(row, 1).text())
        self.price_input.setValue(int(self.table.item(row, 2).text()))
        self.qty_input.setValue(int(self.table.item(row, 3).text()))
    
    def add_bicycle(self):
        """새 자전거 추가"""
        if not self.name_input.text():
            QMessageBox.warning(self, '경고', '자전거명을 입력하세요.')
            return
        
        try:
            self.cursor.execute(
                'INSERT INTO Bycle (name, price, qty) VALUES (?, ?, ?)',
                (self.name_input.text(), self.price_input.value(), self.qty_input.value())
            )
            self.conn.commit()
            QMessageBox.information(self, '성공', '자전거가 추가되었습니다.')
            self.clear_inputs()
            self.load_all_data()
        except Exception as e:
            QMessageBox.critical(self, '오류', f'추가 실패: {str(e)}')
    
    def update_bicycle(self):
        """선택된 자전거 정보 수정"""
        if not self.search_id.text():
            QMessageBox.warning(self, '경고', 'ID를 선택하세요.')
            return
        
        try:
            bicycle_id = int(self.search_id.text())
            self.cursor.execute(
                'UPDATE Bycle SET name=?, price=?, qty=? WHERE id=?',
                (self.name_input.text(), self.price_input.value(), 
                 self.qty_input.value(), bicycle_id)
            )
            self.conn.commit()
            QMessageBox.information(self, '성공', '자전거 정보가 수정되었습니다.')
            self.clear_inputs()
            self.load_all_data()
        except ValueError:
            QMessageBox.warning(self, '경고', '유효한 ID를 선택하세요.')
        except Exception as e:
            QMessageBox.critical(self, '오류', f'수정 실패: {str(e)}')
    
    def delete_bicycle(self):
        """선택된 자전거 삭제"""
        if not self.search_id.text():
            QMessageBox.warning(self, '경고', 'ID를 선택하세요.')
            return
        
        reply = QMessageBox.question(self, '확인', '정말 삭제하시겠습니까?',
                                     QMessageBox.StandardButton.Yes | 
                                     QMessageBox.StandardButton.No)
        
        if reply == QMessageBox.StandardButton.No:
            return
        
        try:
            bicycle_id = int(self.search_id.text())
            self.cursor.execute('DELETE FROM Bycle WHERE id=?', (bicycle_id,))
            self.conn.commit()
            QMessageBox.information(self, '성공', '자전거가 삭제되었습니다.')
            self.clear_inputs()
            self.load_all_data()
        except ValueError:
            QMessageBox.warning(self, '경고', '유효한 ID를 선택하세요.')
        except Exception as e:
            QMessageBox.critical(self, '오류', f'삭제 실패: {str(e)}')
    
    def search_by_id(self):
        """ID로 검색"""
        try:
            if not self.search_id.text():
                self.load_all_data()
                return
            
            bicycle_id = int(self.search_id.text())
            self.cursor.execute('SELECT * FROM Bycle WHERE id=?', (bicycle_id,))
            row = self.cursor.fetchone()
            
            if row:
                self.display_data([row])
            else:
                QMessageBox.information(self, '검색 결과', '해당 ID를 찾을 수 없습니다.')
                self.load_all_data()
        except ValueError:
            QMessageBox.warning(self, '경고', '올바른 ID를 입력하세요.')
    
    def clear_inputs(self):
        """입력 필드 초기화"""
        self.search_id.clear()
        self.name_input.clear()
        self.price_input.setValue(0)
        self.qty_input.setValue(0)
        self.table.clearSelection()
    
    def export_to_excel(self):
        """데이터베이스의 모든 데이터를 엑셀 파일로 내보내기"""
        try:
            # 모든 데이터 조회
            self.cursor.execute('SELECT * FROM Bycle ORDER BY id')
            rows = self.cursor.fetchall()
            
            if not rows:
                QMessageBox.warning(self, '경고', '내보낼 데이터가 없습니다.')
                return
            
            # 워크북 생성
            wb = Workbook()
            ws = wb.active
            ws.title = '자전거목록'
            
            # 헤더 설정
            headers = ['ID', '자전거명', '가격', '수량']
            ws.append(headers)
            
            # 헤더 스타일
            header_fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=12)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 데이터 입력
            for row in rows:
                ws.append(row)
            
            # 열 너비 조정
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 10
            
            # 데이터 셀 정렬 및 스타일
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    if cell.column == 2:  # 자전거명 열은 왼쪽 정렬
                        cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # 파일 저장
            filename = 'bicycle_data.xlsx'
            wb.save(filename)
            
            QMessageBox.information(self, '성공', f'엑셀 파일이 저장되었습니다.\n파일명: {filename}')
        
        except Exception as e:
            QMessageBox.critical(self, '오류', f'엑셀 파일 내보내기 실패: {str(e)}')
    
    def apply_stylesheet(self):
        """애플리케이션 스타일시트 적용"""
        stylesheet = """
        QMainWindow {
            background-color: #f5f5f5;
        }
        
        QLabel {
            color: #333333;
            font-size: 11px;
            font-weight: 500;
        }
        
        QLineEdit {
            border: 2px solid #ddd;
            border-radius: 5px;
            padding: 8px;
            background-color: white;
            color: #333;
            font-size: 11px;
            selection-background-color: #4CAF50;
        }
        
        QLineEdit:focus {
            border: 2px solid #4CAF50;
            background-color: #f9fff9;
        }
        
        QSpinBox {
            border: 2px solid #ddd;
            border-radius: 5px;
            padding: 5px;
            background-color: white;
            color: #333;
            font-size: 11px;
        }
        
        QSpinBox:focus {
            border: 2px solid #4CAF50;
            background-color: #f9fff9;
        }
        
        QPushButton {
            background-color: #4CAF50;
            color: white;
            border: none;
            border-radius: 5px;
            padding: 10px 20px;
            font-size: 11px;
            font-weight: bold;
            min-width: 80px;
        }
        
        QPushButton:hover {
            background-color: #45a049;
            transform: scale(1.02);
        }
        
        QPushButton:pressed {
            background-color: #3d8b40;
        }
        
        QPushButton#deleteBtn {
            background-color: #f44336;
        }
        
        QPushButton#deleteBtn:hover {
            background-color: #da190b;
        }
        
        QTableWidget {
            border: 1px solid #ddd;
            border-radius: 5px;
            background-color: white;
            gridline-color: #e0e0e0;
            font-size: 11px;
        }
        
        QTableWidget::item {
            padding: 5px;
            border: none;
        }
        
        QTableWidget::item:selected {
            background-color: #4CAF50;
            color: white;
        }
        
        QTableWidget::item:hover {
            background-color: #e8f5e9;
        }
        
        QHeaderView::section {
            background-color: #4CAF50;
            color: white;
            padding: 8px;
            border: none;
            font-size: 11px;
            font-weight: bold;
        }
        
        QMessageBox {
            background-color: #f5f5f5;
        }
        
        QMessageBox QLabel {
            color: #333;
        }
        
        QMessageBox QPushButton {
            min-width: 80px;
        }
        """
        self.setStyleSheet(stylesheet)
    
    def closeEvent(self, event):
        """프로그램 종료 시 데이터베이스 연결 종료"""
        self.conn.close()
        event.accept()


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = BicycleManager()
    window.show()
    sys.exit(app.exec())
